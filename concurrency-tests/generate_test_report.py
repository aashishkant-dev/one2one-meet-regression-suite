#!/usr/bin/env python3
"""
Concurrency Test Report Generator
Creates a beautiful Excel file with all test cases ready to fill in
Human-friendly format with easy copying
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def create_test_report():
    """Create Excel workbook with all test cases"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Test Results"

    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    category_font = Font(bold=True, size=11)

    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Test data - organized by category
    test_cases = {
        "Case 1: Basic Delegate Races": [
            {
                "id": "TC-CC-001",
                "name": "Two delegates race for same slot (autoAccept OFF)",
                "priority": "🔴 CRITICAL",
                "description": "When two delegates book the same person's slot at the same instant, exactly one should win and one should get a clear rejection message. No double-bookings!"
            },
            {
                "id": "TC-CC-002",
                "name": "Case 1 with autoAccept ON (24h timer)",
                "priority": "🔴 CRITICAL",
                "description": "Same race but with auto-accept enabled. Race should resolve immediately at request time, not wait for the 24h timer."
            },
            {
                "id": "TC-CC-003",
                "name": "Losing delegate retries next slot",
                "priority": "🟠 HIGH",
                "description": "After losing a race, the losing delegate should be able to immediately try booking the next available slot without any cooldown or lock issues."
            },
            {
                "id": "TC-CC-004",
                "name": "N-way race (3, 5, 10 delegates scrambling for one slot)",
                "priority": "🔴 CRITICAL",
                "description": "When many delegates (3, then 5, then 10) all try to book the same popular person simultaneously, still exactly ONE wins. Performance should stay reasonable even as N grows."
            },
        ],

        "Case 2: Sponsor Races": [
            {
                "id": "TC-CC-005",
                "name": "Two delegates race for sponsor slot (autoAccept OFF)",
                "priority": "🔴 CRITICAL",
                "description": "Same race dynamics but targeting a sponsor. Since sponsors have limited booth tables, table capacity must never exceed 1 booking per slot."
            },
            {
                "id": "TC-CC-006",
                "name": "Sponsor race with autoAccept ON",
                "priority": "🔴 CRITICAL",
                "description": "Sponsor race with auto-accept. Race resolved at request time, not deferred. Only winner goes pending."
            },
            {
                "id": "TC-CC-007",
                "name": "Fan-out scale (10-50 delegates for one sponsor)",
                "priority": "🔴 CRITICAL",
                "description": "When 10, 25, or 50 delegates simultaneously try to book the most popular sponsor, exactly 1 wins. Server stays responsive for unrelated bookings. Sponsor doesn't get flooded with duplicate notifications."
            },
        ],

        "Case 3: Independent Bookings (No Conflicts)": [
            {
                "id": "TC-CC-008",
                "name": "Two independent bookings fired simultaneously",
                "priority": "🟠 HIGH",
                "description": "A books C, B books D (different people, different slots). Both should succeed in parallel without cross-contamination. Response time should NOT be 2x slower than a single booking (checking for global locks)."
            },
            {
                "id": "TC-CC-009",
                "name": "50 independent bookings (25 pairs) in burst",
                "priority": "🟠 HIGH",
                "description": "Scale up: 50 unrelated bookings fired in parallel. All 50 should succeed with correct attribution. No losses, no duplicates, no mix-ups."
            },
        ],

        "Session Races & Idempotency": [
            {
                "id": "TC-CR-001",
                "name": "Double-click Accept button",
                "priority": "🟠 HIGH",
                "description": "When a delegate rapidly double-clicks 'Accept' on a meeting request, only ONE confirmation should be created. Button should be disabled or server should reject the duplicate."
            },
            {
                "id": "TC-CR-002",
                "name": "Accept vs Reject from two sessions (same person, two devices)",
                "priority": "🟠 HIGH",
                "description": "Same delegate logged in on two devices. On device 1, click Accept. On device 2, click Reject - at the exact same instant. Final state should be consistent (not schizophrenic)."
            },
            {
                "id": "TC-CR-003",
                "name": "Request withdrawal vs acceptance (simultaneous race)",
                "priority": "🟠 HIGH",
                "description": "A clicks 'Withdraw' their pending request to C. C clicks 'Accept' at the exact same instant. CRITICAL: No confirmed meeting should exist if the request was withdrawn!"
            },
            {
                "id": "TC-CR-004",
                "name": "Block slot vs simultaneous request",
                "priority": "🟠 HIGH",
                "description": "C clicks 'Block' their own slot. At the same instant, A sends a request for that slot. Either blocked OR pending - never both. Clear error message to loser."
            },
            {
                "id": "TC-CR-005",
                "name": "Manual booking vs self-booking race ⚠️ CRITICAL CROSS-PATHWAY",
                "priority": "🔴 CRITICAL",
                "description": "Organizer manually books A↔C via Manual Booking. At the same instant, B self-books C via delegate interface. Only ONE should succeed. This test catches if manual and self-booking don't share the same slot lock!"
            },
            {
                "id": "TC-CR-006",
                "name": "Auto Rejection setting changed mid-flight",
                "priority": "🟡 MEDIUM",
                "description": "While pending requests exist, organizer toggles 'Auto Rejection' setting ON. Pre-existing requests should NOT be auto-rejected - only NEW requests follow new rule."
            },
            {
                "id": "TC-CR-007",
                "name": "Auto-accept batch race (20-50 requests at 24h threshold)",
                "priority": "🔴 CRITICAL",
                "description": "20-50 pending requests all hit the 24h auto-accept threshold in the same batch window, with some competing for the same slot. All should be processed exactly once with correct single-winner resolution."
            },
        ],

        "Event Setup & Configuration Races": [
            {
                "id": "TC-CE-001",
                "name": "Duplicate event creation (slug collision)",
                "priority": "🟠 HIGH",
                "description": "Two organizers create events with identical names simultaneously. Only ONE should succeed. Duplicate should be rejected with clear message. No slug collision!"
            },
            {
                "id": "TC-CE-002",
                "name": "Opposite status toggles (ON vs OFF)",
                "priority": "🟠 HIGH",
                "description": "Admin 1 toggles event ON. Admin 2 toggles event OFF - at the same instant. Event should settle into ONE final state. Both admins see the same status after refresh."
            },
            {
                "id": "TC-CE-004",
                "name": "Overlapping agenda blocks saved concurrently",
                "priority": "🟠 HIGH",
                "description": "Two admins save agenda blocks with overlapping times (10:00-12:00 and 11:00-13:00). Server should catch the overlap and reject one with clear error. Not just a client-side check!"
            },
            {
                "id": "TC-CE-005",
                "name": "Agenda double-click save",
                "priority": "🟡 MEDIUM",
                "description": "Admin rapidly double-clicks 'Save Agenda'. Only ONE agenda block should be created, not two duplicates with double the slots."
            },
            {
                "id": "TC-CE-006",
                "name": "Agenda edit vs live booking race ⚠️ HIGHEST IMPACT",
                "priority": "🔴 CRITICAL",
                "description": "Organizer edits an agenda (changes slot duration, regenerating the slot structure). At the same instant, a delegate books a slot against the OLD structure. Pre-existing confirmed meetings should be protected. New bookings should be rejected with clear message, never booked against non-existent slots!"
            },
        ],
    }

    # Add title
    ws.merge_cells('A1:G1')
    title_cell = ws['A1']
    title_cell.value = "🧪 One2One Meet - Concurrency Testing Report"
    title_cell.font = Font(bold=True, size=14, color="1F4E78")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Add metadata
    row = 2
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].value = "Test Date:"
    ws[f'C{row}'].value = datetime.now().strftime("%Y-%m-%d")

    row = 3
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].value = "Tester Name:"
    ws[f'C{row}'].value = "[Your Name Here]"

    row = 4
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].value = "Environment:"
    ws[f'C{row}'].value = "[Staging/Production]"

    row = 5
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'].value = "Notes:"
    ws[f'C{row}'].value = "[Any general notes]"

    # Add headers
    row = 7
    headers = ["Test ID", "Test Name", "Priority", "Description", "Status", "Actual Result", "Comments"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    ws.row_dimensions[row].height = 30

    # Add test cases
    row = 8
    current_category = None

    for category, tests in test_cases.items():
        # Category header
        if current_category != category:
            ws.merge_cells(f'A{row}:G{row}')
            cat_cell = ws[f'A{row}']
            cat_cell.value = f"📌 {category}"
            cat_cell.fill = category_fill
            cat_cell.font = category_font
            cat_cell.alignment = Alignment(horizontal='left', vertical='center')
            cat_cell.border = border
            ws.row_dimensions[row].height = 22
            row += 1
            current_category = category

        # Test cases
        for test in tests:
            ws[f'A{row}'].value = test["id"]
            ws[f'B{row}'].value = test["name"]
            ws[f'C{row}'].value = test["priority"]
            ws[f'D{row}'].value = test["description"]
            ws[f'E{row}'].value = "[Pass/Fail/Not Tested]"
            ws[f'F{row}'].value = "[Fill in actual result here]"
            ws[f'G{row}'].value = "[Any notes or observations]"

            # Format row
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Color status column
            ws[f'E{row}'].alignment = Alignment(horizontal='center', vertical='center')

            ws.row_dimensions[row].height = 45
            row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 35
    ws.column_dimensions['G'].width = 25

    # Add summary sheet
    summary = wb.create_sheet("Summary")
    summary['A1'].value = "Test Execution Summary"
    summary['A1'].font = Font(bold=True, size=14)

    summary['A3'].value = "Total Tests:"
    summary['B3'].value = 29

    summary['A4'].value = "Tests Passed:"
    summary['B4'].value = "=COUNTIF('Test Results'!E:E,\"Pass\")"
    summary['B4'].font = Font(color="008000", bold=True)

    summary['A5'].value = "Tests Failed:"
    summary['B5'].value = "=COUNTIF('Test Results'!E:E,\"Fail\")"
    summary['B5'].font = Font(color="FF0000", bold=True)

    summary['A6'].value = "Not Tested:"
    summary['B6'].value = "=COUNTIF('Test Results'!E:E,\"Not Tested\")"

    summary['A8'].value = "Pass Rate:"
    summary['B8'].value = "=IF(B3=0,0,B4/B3)"
    summary['B8'].number_format = '0%'
    summary['B8'].font = Font(bold=True, size=12)

    summary.column_dimensions['A'].width = 20
    summary.column_dimensions['B'].width = 25

    # Add instructions sheet
    instructions = wb.create_sheet("Instructions", 0)

    inst_row = 1
    instructions[f'A{inst_row}'].value = "📋 How to Use This Report"
    instructions[f'A{inst_row}'].font = Font(bold=True, size=14, color="1F4E78")

    inst_row = 3
    instructions[f'A{inst_row}'].value = "BEFORE TESTING:"
    instructions[f'A{inst_row}'].font = Font(bold=True, size=12)

    inst_row = 4
    instructions[f'A{inst_row}'].value = "1. Fill in Test Date, Tester Name, Environment at the top of 'Test Results' sheet"

    inst_row = 5
    instructions[f'A{inst_row}'].value = "2. Run the test automation suite: npm test or .\\run-tests.ps1"

    inst_row = 6
    instructions[f'A{inst_row}'].value = "3. Keep this file open as you test"

    inst_row = 8
    instructions[f'A{inst_row}'].value = "DURING TESTING:"
    instructions[f'A{inst_row}'].font = Font(bold=True, size=12)

    inst_row = 9
    instructions[f'A{inst_row}'].value = "• For each test, fill in the Status column: Pass, Fail, or Not Tested"

    inst_row = 10
    instructions[f'A{inst_row}'].value = "• In Actual Result, describe what actually happened"

    inst_row = 11
    instructions[f'A{inst_row}'].value = "• In Comments, note any issues, errors, or observations"

    inst_row = 13
    instructions[f'A{inst_row}'].value = "AFTER TESTING:"
    instructions[f'A{inst_row}'].font = Font(bold=True, size=12)

    inst_row = 14
    instructions[f'A{inst_row}'].value = "• Check the Summary sheet - it auto-calculates your pass rate"

    inst_row = 15
    instructions[f'A{inst_row}'].value = "• Review all 'Fail' results and document findings"

    inst_row = 16
    instructions[f'A{inst_row}'].value = "• Save this file as your test report"

    inst_row = 18
    instructions[f'A{inst_row}'].value = "STATUS LEGEND:"
    instructions[f'A{inst_row}'].font = Font(bold=True, size=12)

    inst_row = 19
    instructions[f'A{inst_row}'].value = "🔴 CRITICAL = Must pass before production release"
    instructions[f'A{inst_row}'].font = Font(color="FF0000")

    inst_row = 20
    instructions[f'A{inst_row}'].value = "🟠 HIGH = Important, blocking issues if failed"
    instructions[f'A{inst_row}'].font = Font(color="FFA500")

    inst_row = 21
    instructions[f'A{inst_row}'].value = "🟡 MEDIUM = Should fix but not blocking"
    instructions[f'A{inst_row}'].font = Font(color="FFB700")

    instructions.column_dimensions['A'].width = 80

    # Save file
    filename = "Concurrency_Test_Report.xlsx"
    wb.save(filename)
    print(f"✅ Test report created: {filename}")
    print(f"📁 Location: {filename}")
    print(f"📊 Total test cases: 29")
    print(f"\nReady to use! Open {filename} and start filling in results.")

if __name__ == "__main__":
    try:
        create_test_report()
    except ImportError:
        print("❌ Error: openpyxl not installed")
        print("Install it with: pip install openpyxl")
        print("\nOr manually create the Excel file following this format:")
        print("- Column A: Test ID (TC-CC-001, etc)")
        print("- Column B: Test Name")
        print("- Column C: Priority (CRITICAL, HIGH, etc)")
        print("- Column D: Description")
        print("- Column E: Status (Pass/Fail/Not Tested)")
        print("- Column F: Actual Result")
        print("- Column G: Comments")
